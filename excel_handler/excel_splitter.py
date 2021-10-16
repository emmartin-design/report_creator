import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles.colors import Color
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font


from utilities.utility_functions import (
    generate_delimiters,
    check_against_truth_threshold,
    remove_items_from_list
)


question_cell_colors = ['FFFBE4E4', 'FFFBE4E4', 'FFEEEDEC']
non_value_colors = ['00000000', 'FFC0C0C0']

value_storage = ['Sample Size', 'Mentions', '% Sample Size', 'Count', 'Column %', 'Mean', 'Median']

first_row_of_data = 7
column_header_row = 6


def add_formatting(cell, wrap=True, h_align='center', v_align='center', size=8, bold=False, cell_color=None, float=True
                   ):
    cell.alignment = Alignment(horizontal=h_align, vertical=v_align, wrap_text=wrap)
    cell.font = Font(bold=bold, size=size)
    cell.border = Border(top=None, left=None, right=None, bottom=None)
    if cell_color is not None:
        PatternFill(fill_type='solid', fgColor=cell_color)
    try:
        if cell.value < 1:
            cell.number_format = '0%'
        else:
            cell.number_format = '0.0' if float else '0'
    except TypeError:
        pass


def create_mergable_list(lst):
    most_recent_value = None
    new_lst = []
    for x in lst:
        if x is not None:
            new_lst.append(x)
            most_recent_value = x
        else:
            new_lst.append(most_recent_value)
    return new_lst


def move_worksheet(wb, worksheet_to_move, new_position):
    order = [wb.sheetnames.index(x) for x in wb.sheetnames]
    popped = order.pop(wb.sheetnames.index(worksheet_to_move))
    order.insert(new_position, popped)
    wb._sheets = [wb._sheets[i] for i in order]


def create_data_worksheets(file, data_dict):
    with pd.ExcelWriter(file) as writer:
        for idx, (question, info) in enumerate(data_dict.items()):
            sheet_name = f'T{(idx + 1)}'
            data = info['frame']
            data.to_excel(
                writer,
                sheet_name=sheet_name,
                startrow=3,
                index_label= False,
                engine='openpyxl'
            )


def general_format_all_cells(ws, row_start=4):
    for col in ws.iter_cols(min_row=row_start):
        for cell in col:
            add_formatting(cell)


def format_series_headers(ws, row_start=4, row_end=5):
    for col in ws.iter_cols(min_row=row_start, max_row=row_end):
        for cell_idx, cell in enumerate(col):
            add_formatting(cell, h_align='center', v_align='bottom', bold=True if cell_idx < 2 else False)


def format_bases(ws, row_start=6, row_end=6):
    for col in ws.iter_cols(min_row=row_start, max_row=row_end):
        for cell_idx, cell in enumerate(col):
            add_formatting(cell, h_align='center', v_align='center', float=False)


def format_category_headers(ws, is_multiindex=False):
    for col in ws.iter_cols(min_col=1, max_col=2 if is_multiindex else 1):
        for cell_idx, cell in enumerate(col):
            add_formatting(cell, h_align='left', v_align='center', bold=True)


def add_and_format_defaults(ws, question, is_multiindex):
    defaults = {
        'A1': {'val': question},
        'A3': {'val': 'bar'},
        'B3': {'val': '*Sort'},
        'B6' if is_multiindex else 'A6': {'val': 'Base'}
    }
    for idx, (place, contents) in enumerate(defaults.items()):
        ws[place] = contents['val']
        add_formatting(
            ws[place],
            h_align='right' if contents['val'] == 'Base' else 'left',
            wrap=False, bold=(True if idx == 0 else False),
            size=12 if idx == 0 else 8
        )


def adjust_widths(ws, columns, width=20):
    for column in columns:
        ws.column_dimensions[column].width = width


def adjust_heights(ws):
    ws.row_dimensions[4].height = 25
    for row in range(6, (ws.max_row + 1)):
        ws.row_dimensions[row].height = 25


def preselect_data(ws, is_multiindex):
    #  c = Color(indexed=32)
    #  c = Color(theme=6, tint=0.5)
    colors = {
        'A1': Color(theme=4),
        'C6' if is_multiindex else 'B6': Color(theme=5),
        'A8' if is_multiindex else 'B5': Color(theme=7),
    }
    data_selection = Color(theme=7)
    for cell, color in colors.items():
        ws[cell].fill = PatternFill("solid", fgColor=color)
    for col in ws.iter_cols(min_row=8 if is_multiindex else 7, max_col=3 if is_multiindex else 2):
        for cell in col:
            cell.fill = PatternFill("solid", fgColor=data_selection)


def format_data_worksheets(wb, data_dict):
    questions = data_dict.keys()
    for sheet, question in zip(wb.sheetnames, questions):
        is_multiindex = isinstance(data_dict[question]['frame'].index, pd.MultiIndex)

        ws = wb[sheet]
        if not is_multiindex:
            ws.delete_rows(7)
        general_format_all_cells(ws)
        format_series_headers(ws)
        format_bases(ws)
        format_category_headers(ws, is_multiindex)
        add_and_format_defaults(ws, question, is_multiindex)
        adjust_widths(ws, ['A', 'B'] if is_multiindex else ['A'])
        adjust_heights(ws)

        preselect_data(ws, is_multiindex)


def create_contents_page(wb, entries):
    sheet = wb.create_sheet(title='Contents')
    sheet.append(['Contents'])
    for q_idx, question in enumerate(entries):
        sheet.append([f'T{(q_idx + 1)}', question])
    move_worksheet(wb, 'Contents', 0)


def create_new_workbook(file_name, data_dict):
    output_file_name = file_name.replace('.xlsx', '_output.xlsx')
    create_data_worksheets(output_file_name, data_dict)

    wb = load_workbook(output_file_name)
    print(wb.loaded_theme)
    format_data_worksheets(wb, data_dict)
    create_contents_page(wb, data_dict.keys())
    wb.save(output_file_name)


def clean_up_questions(questions, question_indices):
    most_recent_question = questions[0].split()
    reject_questions, reject_indexes = [], []
    for q_idx, (question, question_index) in enumerate(zip(questions, question_indices)):
        if q_idx != 0:
            q_split = question.split()
            similarity_check = [x == y for x, y in zip(most_recent_question, q_split)]

            if check_against_truth_threshold(similarity_check, 0.75):
                reject_questions.append(question_index)
                reject_indexes.append(question_index)

            most_recent_question = q_split
    questions = [x for x in questions if x not in reject_questions]
    question_indices = [x for x in question_indices if x not in reject_indexes]
    return questions, question_indices


def comparative_separation(text, text_to_compare_against):
    text_list = text.split()
    comparative_list = text_to_compare_against.split()

    new_text_list = [x for x_idx, x in enumerate(text_list) if x == comparative_list[x_idx]]
    new_text = ' '.join(new_text_list)
    excluded_text = ' '.join([y for y in text_list if y not in new_text_list])
    excluded_text = None if len(excluded_text) == 1 else excluded_text

    return new_text, excluded_text


def clean_up_question(question):
    outline_delimiters = generate_delimiters(suffix='\t')
    for od in outline_delimiters:
        question = question.replace(od, ': ')
    component_lst = question.split(': ')
    component_lst = [x.strip() for x in component_lst]

    clean_values = {'prefix': component_lst[0], 'question': component_lst[1]}
    try:
        clean_values['statement'] = component_lst[2]
    except IndexError:
        clean_values['statement'] = None

    return clean_values


def clean_up_frame(df, statement=None):
    df = df.set_index(df.columns[0])
    value_row_names = [x for x in value_storage if x in df.index.tolist()]
    value_row_titles = [x for x in value_row_names if any(['%' in x, x in ['Median', 'Mean']])]
    responses = remove_items_from_list(df.index.tolist(), value_row_names)
    responses = value_row_titles if len(responses) == 0 else responses
    df = df[df.index.isin(value_row_titles)]

    if statement is not None:
        statement_array = [statement for _ in responses]
        df['Statement'] = statement_array
        df.index = pd.MultiIndex.from_arrays([statement_array, responses], names=('Statement', 'Response'))
    else:
        df['Responses'] = responses
        df = df.set_index('Responses', drop=True)
        df = df[[x for x in df.columns if x != 'Responses']]

    return df


def get_list_of_fills(ws):
    colors, color_idx = [], []
    for row_idx, row in enumerate(ws.iter_rows(max_col=1)):
        for cell in row:
            color = cell.fill.start_color.rgb
            colors.append(color)
            color_idx = color_idx + ([row_idx] if color in question_cell_colors else [])

    return color_idx


def df_question_scrubber(df):
    index = df.index.tolist()
    split_qs = [x.split() for x in index]
    word_groups = list(zip(*split_qs))

    split_idx = 0
    for group_idx, group in enumerate(word_groups):
        if len(set(group)) != 1:
            split_idx = group_idx
            break

    new_index = [' '.join(x[split_idx:]) for x in split_qs]
    for i, n in zip(index, new_index):
        df =df.rename(index={i: n})

    new_q = ' '.join(split_qs[0][:split_idx])
    new_q = None if len(new_q) == 0 else new_q

    return df, new_q


def data_cleanup(data_dict):
    temp_sheets = {}
    for question, info in data_dict.items():
        try:
            info['frame'], new_q = df_question_scrubber(info['frame'])
        except AttributeError:
            new_q = None
        if new_q is None:
            temp_sheets[question] = info
        else:
            temp_sheets[new_q] = info

    return temp_sheets


def split_excel(file_name):
    """
    Splits original df in to new dataframes to be shuffled and reorganized later
    """
    og_wb = load_workbook(filename=file_name)

    new_sheets = {}
    for sheet in og_wb.worksheets:
        headers = ['Question Values'] + [x.value for x in sheet[column_header_row]][1:]

        cat_headers = create_mergable_list([x.value for x in sheet[column_header_row - 1]][1:])
        cat_headers = [x.split(': ') for x in cat_headers]
        cat_headers = ['Question Values'] + [x[1] for x in cat_headers]
        df = pd.DataFrame(sheet.values)
        sample_size_indices = df.index[df[0] == 'Sample Size'].tolist()
        sample_sizes = df.loc[sample_size_indices[0]].values.tolist()

        question_indices = get_list_of_fills(sheet)
        questions = [x[0] for x in df.filter(items=question_indices, axis=0).values.tolist()]

        df = df.iloc[first_row_of_data:]
        df.columns = pd.MultiIndex.from_arrays([cat_headers, headers, sample_sizes])

        for index_idx, index in enumerate(question_indices):
            question_values = clean_up_question(questions[index_idx])
            # most_recent_question = questions[index_idx]

            if index_idx == len(question_indices) - 1:
                q_df = df.loc[(index + 1):]
            else:
                next_index = (question_indices[index_idx + 1] - 1)
                q_df = df.loc[(index + 1):next_index]

            q_df = clean_up_frame(q_df, question_values['statement'])

            if new_sheets.get(question_values['question']):
                current_frame = new_sheets[question_values['question']]['frame']
                new_sheets[question_values['question']]['frame'] = current_frame.append(q_df)
            else:
                new_sheets[question_values['question']] = {'frame': q_df, 'sample': sample_sizes}

    new_sheets = data_cleanup(new_sheets)
    create_new_workbook(file_name, new_sheets)
